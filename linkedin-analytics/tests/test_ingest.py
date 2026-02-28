"""Tests for the LinkedIn analytics ingestion pipeline."""

import hashlib
import io
from datetime import date
from pathlib import Path

import openpyxl
import pytest

from app.ingest import (
    DuplicateFileError,
    ImportStats,
    IngestError,
    ParsedExport,
    compute_file_hash,
    ingest_file,
    load_to_db,
    parse_linkedin_export,
    validate_upload,
)
from app.models import DailyMetric, DemographicSnapshot, FollowerSnapshot, Post, Upload


# ---------------------------------------------------------------------------
# compute_file_hash
# ---------------------------------------------------------------------------


class TestComputeFileHash:
    def test_returns_sha256_hex(self, tmp_path):
        f = tmp_path / "test.bin"
        f.write_bytes(b"hello world")
        result = compute_file_hash(f)
        expected = hashlib.sha256(b"hello world").hexdigest()
        assert result == expected

    def test_different_content_different_hash(self, tmp_path):
        f1 = tmp_path / "a.bin"
        f2 = tmp_path / "b.bin"
        f1.write_bytes(b"content a")
        f2.write_bytes(b"content b")
        assert compute_file_hash(f1) != compute_file_hash(f2)

    def test_same_content_same_hash(self, tmp_path):
        f1 = tmp_path / "a.bin"
        f2 = tmp_path / "b.bin"
        f1.write_bytes(b"same content")
        f2.write_bytes(b"same content")
        assert compute_file_hash(f1) == compute_file_hash(f2)


# ---------------------------------------------------------------------------
# validate_upload
# ---------------------------------------------------------------------------


class TestValidateUpload:
    def test_valid_xlsx(self, sample_xlsx_path):
        validate_upload(sample_xlsx_path)  # Should not raise

    def test_file_not_found(self, tmp_path):
        with pytest.raises(IngestError, match="not found"):
            validate_upload(tmp_path / "nonexistent.xlsx")

    def test_empty_file(self, tmp_path):
        f = tmp_path / "empty.xlsx"
        f.write_bytes(b"")
        with pytest.raises(IngestError, match="empty"):
            validate_upload(f)

    def test_unsupported_extension(self, tmp_path):
        f = tmp_path / "data.txt"
        f.write_bytes(b"some content")
        with pytest.raises(IngestError, match="Unsupported file type"):
            validate_upload(f)

    def test_csv_allowed(self, tmp_path):
        f = tmp_path / "data.csv"
        f.write_text("Date,Impressions\n2025-11-01,100\n")
        validate_upload(f)  # Should not raise


# ---------------------------------------------------------------------------
# parse_linkedin_export
# ---------------------------------------------------------------------------


class TestParseLinkedinExport:
    def test_parse_sample_xlsx(self, sample_xlsx_path):
        result = parse_linkedin_export(sample_xlsx_path)
        assert isinstance(result, ParsedExport)
        assert len(result.posts) >= 5
        assert len(result.daily_metrics) > 0
        assert len(result.follower_snapshots) > 0
        assert len(result.demographic_snapshots) > 0

    def test_posts_have_required_fields(self, sample_xlsx_path):
        result = parse_linkedin_export(sample_xlsx_path)
        for post in result.posts:
            assert "post_date" in post
            assert isinstance(post["post_date"], date)
            assert post.get("impressions", 0) >= 0

    def test_engagement_rate_calculated(self, sample_xlsx_path):
        result = parse_linkedin_export(sample_xlsx_path)
        for post in result.posts:
            impressions = post.get("impressions", 0)
            if impressions > 0:
                expected = (
                    post.get("reactions", 0)
                    + post.get("comments", 0)
                    + post.get("shares", 0)
                ) / impressions
                assert post["engagement_rate"] == pytest.approx(expected, rel=1e-3)

    def test_missing_discovery_sheet_warns_not_crashes(self, tmp_path):
        """A file without the DISCOVERY sheet should warn but still parse."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ENGAGEMENT"
        ws.append(["Post Date", "Post Title", "Impressions", "Reactions", "Comments", "Shares", "Clicks"])
        ws.append(["2025-11-01", "Test post", 1000, 50, 10, 5, 20])
        out = tmp_path / "partial.xlsx"
        wb.save(out)

        result = parse_linkedin_export(out)
        assert any("DISCOVERY" in w for w in result.warnings)
        assert len(result.posts) == 1

    def test_missing_followers_sheet_warns_not_crashes(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "DISCOVERY"
        ws.append(["Date", "Impressions"])
        ws.append(["2025-11-01", 200])
        out = tmp_path / "no_followers.xlsx"
        wb.save(out)

        result = parse_linkedin_export(out)
        assert any("FOLLOWERS" in w for w in result.warnings)

    def test_malformed_data_rows_skipped(self, tmp_path):
        """Rows with unparseable dates should be skipped without crashing."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ENGAGEMENT"
        ws.append(["Post Date", "Post Title", "Impressions", "Reactions", "Comments", "Shares", "Clicks"])
        ws.append(["NOT_A_DATE", "Bad row", "abc", "xyz", "", "", ""])
        ws.append(["2025-11-01", "Good row", 1000, 50, 10, 5, 20])
        out = tmp_path / "malformed.xlsx"
        wb.save(out)

        result = parse_linkedin_export(out)
        assert len(result.posts) == 1
        assert result.posts[0]["title"] == "Good row"

    def test_empty_file_raises(self, tmp_path):
        f = tmp_path / "empty.xlsx"
        f.write_bytes(b"")
        with pytest.raises(IngestError):
            parse_linkedin_export(f)

    def test_unsupported_extension_raises(self, tmp_path):
        f = tmp_path / "data.txt"
        f.write_bytes(b"some data")
        with pytest.raises(IngestError):
            parse_linkedin_export(f)

    def test_csv_file_parsed(self, tmp_path):
        """A CSV file should be accepted and treated as a single sheet."""
        f = tmp_path / "DISCOVERY.csv"
        f.write_text("Date,Impressions,Members Reached\n2025-11-01,200,140\n2025-11-02,250,175\n")
        result = parse_linkedin_export(f)
        # CSV doesn't produce posts but should not crash
        assert isinstance(result, ParsedExport)

    def test_title_truncated_to_100_chars(self, tmp_path):
        long_title = "A" * 200
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ENGAGEMENT"
        ws.append(["Post Date", "Post Title", "Impressions", "Reactions", "Comments", "Shares", "Clicks"])
        ws.append(["2025-11-01", long_title, 1000, 50, 10, 5, 20])
        out = tmp_path / "longtitle.xlsx"
        wb.save(out)

        result = parse_linkedin_export(out)
        assert len(result.posts[0]["title"]) <= 100


# ---------------------------------------------------------------------------
# load_to_db
# ---------------------------------------------------------------------------


class TestLoadToDb:
    def test_load_posts(self, test_session):
        parsed = ParsedExport(
            posts=[
                {
                    "post_date": date(2025, 11, 1),
                    "title": "Test post",
                    "post_type": "text",
                    "impressions": 1000,
                    "members_reached": 800,
                    "reactions": 50,
                    "comments": 10,
                    "shares": 5,
                    "clicks": 20,
                    "engagement_rate": 0.065,
                    "linkedin_post_id": None,
                }
            ]
        )
        stats = load_to_db(test_session, parsed)
        assert stats.posts_upserted == 1
        assert test_session.query(Post).count() == 1

    def test_upsert_higher_value_wins(self, test_session):
        """Re-importing the same post with higher metrics should update the record."""
        parsed1 = ParsedExport(
            posts=[
                {
                    "post_date": date(2025, 11, 1),
                    "title": "Test post",
                    "impressions": 1000,
                    "reactions": 50,
                    "comments": 10,
                    "shares": 5,
                    "clicks": 20,
                    "members_reached": 800,
                    "engagement_rate": 0.065,
                    "linkedin_post_id": None,
                    "post_type": None,
                }
            ]
        )
        load_to_db(test_session, parsed1)

        parsed2 = ParsedExport(
            posts=[
                {
                    "post_date": date(2025, 11, 1),
                    "title": "Test post",
                    "impressions": 1500,  # Higher
                    "reactions": 75,      # Higher
                    "comments": 10,
                    "shares": 5,
                    "clicks": 20,
                    "members_reached": 800,
                    "engagement_rate": 0.0,
                    "linkedin_post_id": None,
                    "post_type": None,
                }
            ]
        )
        load_to_db(test_session, parsed2)

        assert test_session.query(Post).count() == 1
        post = test_session.query(Post).first()
        assert post.impressions == 1500
        assert post.reactions == 75

    def test_upsert_lower_value_not_overwrite(self, test_session):
        """Re-importing with lower metrics should not overwrite."""
        parsed1 = ParsedExport(
            posts=[
                {
                    "post_date": date(2025, 11, 1),
                    "title": "Test post",
                    "impressions": 1000,
                    "reactions": 50,
                    "comments": 10,
                    "shares": 5,
                    "clicks": 20,
                    "members_reached": 800,
                    "engagement_rate": 0.065,
                    "linkedin_post_id": None,
                    "post_type": None,
                }
            ]
        )
        load_to_db(test_session, parsed1)

        parsed2 = ParsedExport(
            posts=[
                {
                    "post_date": date(2025, 11, 1),
                    "title": "Test post",
                    "impressions": 500,  # Lower
                    "reactions": 20,     # Lower
                    "comments": 5,
                    "shares": 2,
                    "clicks": 10,
                    "members_reached": 400,
                    "engagement_rate": 0.054,
                    "linkedin_post_id": None,
                    "post_type": None,
                }
            ]
        )
        load_to_db(test_session, parsed2)

        post = test_session.query(Post).first()
        assert post.impressions == 1000  # Original higher value preserved
        assert post.reactions == 50

    def test_load_full_export(self, test_session, sample_xlsx_path):
        result = parse_linkedin_export(sample_xlsx_path)
        stats = load_to_db(test_session, result)
        assert stats.posts_upserted >= 5
        assert stats.follower_snapshots_upserted >= 30
        assert stats.demographic_snapshots_upserted >= 9

    def test_load_follower_snapshots(self, test_session):
        parsed = ParsedExport(
            follower_snapshots=[
                {"snapshot_date": date(2025, 11, 1), "total_followers": 500, "new_followers": 5},
                {"snapshot_date": date(2025, 11, 2), "total_followers": 505, "new_followers": 5},
            ]
        )
        stats = load_to_db(test_session, parsed)
        assert stats.follower_snapshots_upserted == 2
        assert test_session.query(FollowerSnapshot).count() == 2

    def test_load_demographic_snapshots(self, test_session):
        parsed = ParsedExport(
            demographic_snapshots=[
                {"snapshot_date": date(2025, 11, 30), "category": "industry", "value": "IT", "percentage": 30.0},
                {"snapshot_date": date(2025, 11, 30), "category": "industry", "value": "Finance", "percentage": 20.0},
            ]
        )
        stats = load_to_db(test_session, parsed)
        assert stats.demographic_snapshots_upserted == 2


# ---------------------------------------------------------------------------
# ingest_file (full pipeline)
# ---------------------------------------------------------------------------


class TestIngestFile:
    def test_successful_ingest(self, test_session, sample_xlsx_path, tmp_path):
        import shutil
        dest = tmp_path / "export.xlsx"
        shutil.copy(sample_xlsx_path, dest)
        upload, stats = ingest_file(test_session, dest, "export.xlsx")
        assert upload.id is not None
        assert upload.status == "completed"
        assert stats.total_records > 0

    def test_duplicate_file_raises(self, test_session, sample_xlsx_path, tmp_path):
        import shutil
        dest = tmp_path / "export.xlsx"
        shutil.copy(sample_xlsx_path, dest)
        ingest_file(test_session, dest, "export.xlsx")

        # Re-importing the same file should raise DuplicateFileError
        dest2 = tmp_path / "export_copy.xlsx"
        shutil.copy(sample_xlsx_path, dest2)
        with pytest.raises(DuplicateFileError):
            ingest_file(test_session, dest2, "export_copy.xlsx")

    def test_upload_record_created(self, test_session, sample_xlsx_path, tmp_path):
        import shutil
        dest = tmp_path / "export.xlsx"
        shutil.copy(sample_xlsx_path, dest)
        upload, stats = ingest_file(test_session, dest, "export.xlsx")
        db_upload = test_session.query(Upload).filter_by(id=upload.id).first()
        assert db_upload is not None
        assert db_upload.filename == "export.xlsx"
        assert db_upload.records_imported == stats.total_records
