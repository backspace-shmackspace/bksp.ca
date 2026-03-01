"""Tests for per-post XLSX import functionality."""

from datetime import date
from io import BytesIO

import openpyxl
import pytest

from app.ingest import (
    _detect_xlsx_format,
    _extract_urn_from_url,
    _parse_int_with_commas,
    _parse_per_post_demographics,
    _parse_per_post_performance,
    _parse_post_hour,
    ingest_per_post_xlsx,
)
from app.models import Post, PostDemographic


# ---------------------------------------------------------------------------
# Workbook builders
# ---------------------------------------------------------------------------


def _build_per_post_workbook(
    post_url: str = "https://www.linkedin.com/feed/update/urn:li:share:7432391508978397184",
    post_date: str = "Feb 25, 2026",
    post_time: str = "11:53 AM",
    impressions: str = "1,316",
    members_reached: str = "940",
    reactions: str = "42",
    comments: str = "7",
    reposts: str = "3",
    saves: str = "12",
    sends: str = "18",
    profile_views: str = "5",
    followers_gained: str = "2",
    demo_rows: list | None = None,
) -> openpyxl.Workbook:
    """Build a synthetic per-post XLSX workbook."""
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    # PERFORMANCE sheet (key-value layout)
    ws_perf = wb.create_sheet("PERFORMANCE")
    ws_perf.append(["Post URL", post_url])
    ws_perf.append(["Post Date", post_date])
    ws_perf.append(["Post Publish Time", post_time])
    ws_perf.append([None, None])  # spacer
    ws_perf.append(["Impressions", impressions])
    ws_perf.append(["Members reached", members_reached])
    ws_perf.append(["Reactions", reactions])
    ws_perf.append(["Comments", comments])
    ws_perf.append(["Reposts", reposts])
    ws_perf.append(["Saves", saves])
    ws_perf.append(["Sends on LinkedIn", sends])
    ws_perf.append(["Profile viewers from this post", profile_views])
    ws_perf.append(["Followers gained from this post", followers_gained])

    # TOP DEMOGRAPHICS sheet (tabular)
    ws_demo = wb.create_sheet("TOP DEMOGRAPHICS")
    ws_demo.append(["Category", "Value", "Percentage"])
    if demo_rows is None:
        demo_rows = [
            ("Company size", "10,001+ employees", 0.31),
            ("Company size", "1001-5000 employees", 0.18),
            ("Job title", "Security Engineer", 0.22),
            ("Job title", "Software Engineer", 0.15),
            ("Location", "Fredericton", 0.12),
            ("Location", "Greater Toronto Area, Canada", 0.09),
            ("Company", "IBM", 0.08),
            ("Company", "Red Hat", 0.05),
        ]
    for category, value, pct in demo_rows:
        ws_demo.append([category, value, pct])

    return wb


def _build_aggregate_workbook() -> openpyxl.Workbook:
    """Build a minimal aggregate XLSX workbook."""
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    wb.create_sheet("DISCOVERY")
    wb.create_sheet("ENGAGEMENT")
    return wb


# ---------------------------------------------------------------------------
# Format detection
# ---------------------------------------------------------------------------


def test_detect_per_post_format():
    wb = _build_per_post_workbook()
    assert _detect_xlsx_format(wb) == "per_post"


def test_detect_aggregate_format():
    wb = _build_aggregate_workbook()
    assert _detect_xlsx_format(wb) == "aggregate"


def test_detect_unknown_format():
    wb = openpyxl.Workbook()
    wb.active.title = "RANDOM SHEET"
    assert _detect_xlsx_format(wb) == "unknown"


# ---------------------------------------------------------------------------
# Performance sheet parsing
# ---------------------------------------------------------------------------


def test_parse_per_post_performance():
    wb = _build_per_post_workbook(
        post_url="https://www.linkedin.com/feed/update/urn:li:share:1234",
        impressions="1,316",
        reactions="42",
    )
    perf = _parse_per_post_performance(wb["PERFORMANCE"])
    assert perf["Post URL"] == "https://www.linkedin.com/feed/update/urn:li:share:1234"
    assert perf["Impressions"] == "1,316"
    assert perf["Reactions"] == "42"


def test_parse_int_with_commas():
    assert _parse_int_with_commas("1,316") == 1316
    assert _parse_int_with_commas("42") == 42
    assert _parse_int_with_commas("0") == 0
    assert _parse_int_with_commas("bad") == 0
    assert _parse_int_with_commas("") == 0


# ---------------------------------------------------------------------------
# Post hour parsing
# ---------------------------------------------------------------------------


def test_parse_post_hour_am():
    assert _parse_post_hour("11:53 AM") == 11


def test_parse_post_hour_pm():
    assert _parse_post_hour("2:30 PM") == 14


def test_parse_post_hour_midnight():
    assert _parse_post_hour("12:00 AM") == 0


def test_parse_post_hour_noon():
    assert _parse_post_hour("12:00 PM") == 12


def test_parse_post_hour_invalid():
    assert _parse_post_hour("not-a-time") is None
    assert _parse_post_hour("") is None


# ---------------------------------------------------------------------------
# URN extraction
# ---------------------------------------------------------------------------


def test_extract_urn_from_url_share():
    url = "https://www.linkedin.com/feed/update/urn:li:share:7432391508978397184"
    assert _extract_urn_from_url(url) == "7432391508978397184"


def test_extract_urn_from_url_activity():
    url = "https://www.linkedin.com/feed/update/urn:li:activity:6844785523593134080"
    assert _extract_urn_from_url(url) == "6844785523593134080"


def test_extract_urn_from_url_invalid():
    assert _extract_urn_from_url("https://www.linkedin.com/") is None
    assert _extract_urn_from_url("") is None


# ---------------------------------------------------------------------------
# Demographics sheet parsing
# ---------------------------------------------------------------------------


def test_parse_per_post_demographics():
    wb = _build_per_post_workbook()
    rows = _parse_per_post_demographics(wb["TOP DEMOGRAPHICS"])
    assert len(rows) > 0
    categories = {r["category"] for r in rows}
    assert "company_size" in categories or "job_title" in categories


def test_parse_demographics_less_than_one_percent():
    wb = _build_per_post_workbook(
        demo_rows=[
            ("Company size", "Small company", "< 1%"),
        ]
    )
    rows = _parse_per_post_demographics(wb["TOP DEMOGRAPHICS"])
    assert len(rows) == 1
    assert rows[0]["percentage"] == 0.005


def test_parse_demographics_float_percentage():
    wb = _build_per_post_workbook(
        demo_rows=[
            ("Job title", "Director", 0.22),
        ]
    )
    rows = _parse_per_post_demographics(wb["TOP DEMOGRAPHICS"])
    assert rows[0]["percentage"] == pytest.approx(0.22)


def test_parse_demographics_string_percentage():
    wb = _build_per_post_workbook(
        demo_rows=[
            ("Location", "Canada", "15%"),
        ]
    )
    rows = _parse_per_post_demographics(wb["TOP DEMOGRAPHICS"])
    assert rows[0]["percentage"] == pytest.approx(0.15)


# ---------------------------------------------------------------------------
# Full per-post ingest
# ---------------------------------------------------------------------------


def test_ingest_per_post_creates_post(test_session):
    """Ingesting a per-post XLSX creates a Post row with metrics."""
    wb = _build_per_post_workbook(
        post_url="https://www.linkedin.com/feed/update/urn:li:share:1111222233334444",
        impressions="500",
        reactions="15",
        saves="8",
    )
    result = ingest_per_post_xlsx(test_session, wb)

    assert result["linkedin_post_id"] == "1111222233334444"
    assert result["metrics_updated"] is True

    post = test_session.query(Post).filter(Post.id == result["post_id"]).first()
    assert post is not None
    assert post.impressions == 500
    assert post.reactions == 15
    assert post.saves == 8
    assert post.linkedin_post_id == "1111222233334444"


def test_ingest_per_post_updates_existing(test_session):
    """Per-post XLSX updates an existing post matched by linkedin_post_id."""
    existing = Post(
        post_date=date(2026, 2, 25),
        title="My existing post",
        linkedin_post_id="5555666677778888",
        impressions=100,
        reactions=5,
    )
    existing.recalculate_engagement_rate()
    test_session.add(existing)
    test_session.commit()
    test_session.refresh(existing)

    wb = _build_per_post_workbook(
        post_url="https://www.linkedin.com/feed/update/urn:li:share:5555666677778888",
        impressions="1316",
        reactions="42",
        saves="12",
    )
    result = ingest_per_post_xlsx(test_session, wb)
    assert result["post_id"] == existing.id

    test_session.refresh(existing)
    assert existing.impressions == 1316
    assert existing.reactions == 42
    assert existing.saves == 12


def test_ingest_per_post_extracts_new_metrics(test_session):
    """Per-post XLSX stores saves, sends, profile_views, followers_gained, reposts."""
    wb = _build_per_post_workbook(
        post_url="https://www.linkedin.com/feed/update/urn:li:share:9876543210",
        saves="15",
        sends="22",
        profile_views="7",
        followers_gained="3",
        reposts="4",
    )
    result = ingest_per_post_xlsx(test_session, wb)

    post = test_session.query(Post).filter(Post.id == result["post_id"]).first()
    assert post.saves == 15
    assert post.sends == 22
    assert post.profile_views == 7
    assert post.followers_gained == 3
    assert post.reposts == 4


def test_ingest_per_post_extracts_post_hour(test_session):
    """Per-post XLSX stores post_hour in 24-hour format."""
    wb = _build_per_post_workbook(
        post_url="https://www.linkedin.com/feed/update/urn:li:share:1357924680",
        post_time="2:30 PM",
    )
    result = ingest_per_post_xlsx(test_session, wb)

    post = test_session.query(Post).filter(Post.id == result["post_id"]).first()
    assert post.post_hour == 14


def test_per_post_demographics_stored_correctly(test_session):
    """After per-post XLSX import, PostDemographic rows have correct data."""
    wb = _build_per_post_workbook(
        post_url="https://www.linkedin.com/feed/update/urn:li:share:2468135790",
        demo_rows=[
            ("Company size", "10,001+ employees", 0.31),
            ("Job title", "Security Engineer", 0.22),
        ],
    )
    result = ingest_per_post_xlsx(test_session, wb)
    post_id = result["post_id"]

    demos = (
        test_session.query(PostDemographic)
        .filter(PostDemographic.post_id == post_id)
        .all()
    )
    assert len(demos) == 2

    categories = {d.category for d in demos}
    assert "company_size" in categories
    assert "job_title" in categories

    comp_demo = next(d for d in demos if d.category == "company_size")
    assert comp_demo.value == "10,001+ employees"
    assert comp_demo.percentage == pytest.approx(0.31)


def test_ingest_per_post_transitions_status_to_analytics_linked(test_session):
    """When a published post with content gets per-post data, status transitions."""
    existing = Post(
        post_date=date(2026, 2, 25),
        title="Published post",
        content="My post content here.",
        status="published",
        linkedin_post_id="1122334455667788",
    )
    existing.recalculate_engagement_rate()
    test_session.add(existing)
    test_session.commit()
    test_session.refresh(existing)

    wb = _build_per_post_workbook(
        post_url="https://www.linkedin.com/feed/update/urn:li:share:1122334455667788",
    )
    ingest_per_post_xlsx(test_session, wb)

    test_session.refresh(existing)
    assert existing.status == "analytics_linked"


def test_ingest_per_post_demographics_upserted_on_reimport(test_session):
    """Re-importing the same per-post XLSX updates existing demographic rows."""
    wb = _build_per_post_workbook(
        post_url="https://www.linkedin.com/feed/update/urn:li:share:9988776655443322",
        demo_rows=[("Company size", "10,001+ employees", 0.31)],
    )
    result1 = ingest_per_post_xlsx(test_session, wb)
    post_id = result1["post_id"]

    # Re-import with updated percentage
    wb2 = _build_per_post_workbook(
        post_url="https://www.linkedin.com/feed/update/urn:li:share:9988776655443322",
        demo_rows=[("Company size", "10,001+ employees", 0.45)],
    )
    ingest_per_post_xlsx(test_session, wb2)

    demos = (
        test_session.query(PostDemographic)
        .filter(
            PostDemographic.post_id == post_id,
            PostDemographic.category == "company_size",
        )
        .all()
    )
    assert len(demos) == 1
    assert demos[0].percentage == pytest.approx(0.45)


# ---------------------------------------------------------------------------
# Integration: XLSX import linkage
# ---------------------------------------------------------------------------


def test_xlsx_import_preserves_content(test_session):
    """After per-post XLSX import, post content is unchanged."""
    existing = Post(
        post_date=date(2026, 2, 25),
        title="Content post",
        content="Original authored content. Should not be overwritten.",
        status="published",
        linkedin_post_id="3344556677889900",
    )
    existing.recalculate_engagement_rate()
    test_session.add(existing)
    test_session.commit()
    test_session.refresh(existing)

    wb = _build_per_post_workbook(
        post_url="https://www.linkedin.com/feed/update/urn:li:share:3344556677889900",
    )
    ingest_per_post_xlsx(test_session, wb)

    test_session.refresh(existing)
    assert existing.content == "Original authored content. Should not be overwritten."


def test_batch_upload_endpoint(client, tmp_path):
    """POST /api/upload/batch with per-post files returns results."""
    from io import BytesIO
    wb = _build_per_post_workbook(
        post_url="https://www.linkedin.com/feed/update/urn:li:share:5544332211009988",
    )
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = client.post(
        "/api/upload/batch",
        files=[("files", ("post_analytics.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))],
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data["total"] == 1
    assert data["succeeded"] == 1
    assert data["results"][0]["status"] == "ok"
