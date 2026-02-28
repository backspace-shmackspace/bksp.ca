"""Pytest configuration and shared fixtures."""

import io
from datetime import date, timedelta
from pathlib import Path

import openpyxl
import pytest
from fastapi.testclient import TestClient
from sqlalchemy import create_engine, event
from sqlalchemy.orm import sessionmaker, Session

from app.database import get_session, init_db
from app.models import Base, DailyMetric, DemographicSnapshot, FollowerSnapshot, Post, Upload

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

FIXTURES_DIR = Path(__file__).parent / "fixtures"


# ---------------------------------------------------------------------------
# In-memory database fixtures
#
# We use a single shared SQLite connection for the entire test function so that
# data written by test fixtures and data read by the FastAPI route handlers
# both see the same state. SQLite :memory: databases are per-connection and
# data does not propagate across separate connections.
# ---------------------------------------------------------------------------


@pytest.fixture(scope="function")
def test_engine():
    """Create a fresh in-memory SQLite engine per test function.

    Uses a single shared connection so all sessions within a test see the
    same in-memory database state.
    """
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
    )

    # Force all sessions to reuse the same underlying connection
    @event.listens_for(engine, "connect")
    def _enable_fk(dbapi_conn, _):
        cursor = dbapi_conn.cursor()
        cursor.execute("PRAGMA foreign_keys=ON")
        cursor.close()

    Base.metadata.create_all(bind=engine)
    yield engine
    Base.metadata.drop_all(bind=engine)
    engine.dispose()


@pytest.fixture(scope="function")
def test_session(test_engine):
    """Yield a SQLAlchemy session backed by the in-memory database."""
    TestSession = sessionmaker(autocommit=False, autoflush=False, bind=test_engine)
    session = TestSession()
    try:
        yield session
    finally:
        session.close()


# ---------------------------------------------------------------------------
# FastAPI test client
# ---------------------------------------------------------------------------


@pytest.fixture(scope="function")
def client(test_engine, tmp_path):
    """Return a FastAPI TestClient with an isolated in-memory database.

    Both the test session (used to seed data) and the route handler sessions
    share the same SQLite connection via test_engine, so data committed in
    a test helper is visible to the routes.
    """
    from app.main import app
    from app import database as app_db
    from app import config as app_config

    # Use a single shared connection for all sessions in this test
    shared_connection = test_engine.connect()

    TestSession = sessionmaker(
        autocommit=False, autoflush=False, bind=shared_connection
    )

    def override_get_session():
        session = TestSession()
        try:
            yield session
        finally:
            session.close()

    app.dependency_overrides[get_session] = override_get_session

    # Override data_dir so uploads land in tmp_path.
    # We mutate the pydantic field value on the shared settings singleton
    # in place so that all modules that imported the same object (e.g.
    # app.database) see the updated value. uploads_dir and db_path are
    # @property methods that derive from data_dir, so they automatically
    # reflect the override — no separate field mutation is needed.
    uploads_dir = tmp_path / "uploads"
    uploads_dir.mkdir(parents=True, exist_ok=True)

    original_data_dir = app_config.settings.__dict__["data_dir"]
    app_config.settings.__dict__["data_dir"] = tmp_path

    # Patch the global engine used by init_db() so startup uses the test engine
    original_engine = app_db.engine
    original_session_local = app_db.SessionLocal
    app_db.engine = test_engine
    app_db.SessionLocal = TestSession

    with TestClient(app, raise_server_exceptions=True) as c:
        yield c

    app.dependency_overrides.clear()
    shared_connection.close()
    app_db.engine = original_engine
    app_db.SessionLocal = original_session_local
    app_config.settings.__dict__["data_dir"] = original_data_dir


# ---------------------------------------------------------------------------
# Sample database state
# ---------------------------------------------------------------------------


@pytest.fixture
def sample_posts(test_session) -> list[Post]:
    """Insert 5 sample posts into the test database."""
    posts = []
    base_date = date(2025, 11, 1)
    for i in range(5):
        post = Post(
            post_date=base_date + timedelta(days=i * 7),
            title=f"Sample post number {i + 1} about cybersecurity topics"[:100],
            post_type="text",
            impressions=1000 + i * 200,
            members_reached=800 + i * 150,
            reactions=50 + i * 10,
            comments=10 + i * 3,
            shares=5 + i * 2,
            clicks=20 + i * 5,
        )
        post.recalculate_engagement_rate()
        test_session.add(post)
        posts.append(post)
    test_session.commit()
    for p in posts:
        test_session.refresh(p)
    return posts


@pytest.fixture
def sample_follower_snapshots(test_session) -> list[FollowerSnapshot]:
    """Insert 30 daily follower snapshots into the test database."""
    snapshots = []
    base_date = date(2025, 11, 1)
    followers = 450
    for i in range(30):
        new = i % 5
        followers += new
        snapshot = FollowerSnapshot(
            snapshot_date=base_date + timedelta(days=i),
            total_followers=followers,
            new_followers=new,
        )
        test_session.add(snapshot)
        snapshots.append(snapshot)
    test_session.commit()
    return snapshots


@pytest.fixture
def sample_demographics(test_session) -> list[DemographicSnapshot]:
    """Insert sample demographic records into the test database."""
    snap_date = date(2025, 11, 30)
    records = [
        DemographicSnapshot(snapshot_date=snap_date, category="industry", value="Information Technology", percentage=32.5),
        DemographicSnapshot(snapshot_date=snap_date, category="industry", value="Financial Services", percentage=18.2),
        DemographicSnapshot(snapshot_date=snap_date, category="industry", value="Cybersecurity", percentage=14.0),
        DemographicSnapshot(snapshot_date=snap_date, category="job_title", value="Director", percentage=22.0),
        DemographicSnapshot(snapshot_date=snap_date, category="job_title", value="Manager", percentage=18.5),
        DemographicSnapshot(snapshot_date=snap_date, category="seniority", value="Senior", percentage=40.0),
        DemographicSnapshot(snapshot_date=snap_date, category="seniority", value="Director", percentage=25.0),
        DemographicSnapshot(snapshot_date=snap_date, category="location", value="United States", percentage=55.0),
        DemographicSnapshot(snapshot_date=snap_date, category="location", value="Canada", percentage=12.0),
    ]
    for r in records:
        test_session.add(r)
    test_session.commit()
    return records


# ---------------------------------------------------------------------------
# Sample XLS export fixture (synthetic .xlsx created programmatically)
# ---------------------------------------------------------------------------


def _build_sample_xlsx(output_path: Path) -> None:
    """Create a synthetic LinkedIn analytics .xlsx export file for testing.

    Mirrors the real LinkedIn Premium export format (verified Feb 2026):
    - DISCOVERY: summary metrics (label/value pairs)
    - ENGAGEMENT: daily totals (Date | Impressions | Engagements)
    - TOP POSTS: two side-by-side tables (Engagements A-C, Impressions E-G)
    - FOLLOWERS: total at top, then Date | New followers (header at row 3)
    - DEMOGRAPHICS: Category | Value | Percentage (header: "Top Demographics")
    """
    wb = openpyxl.Workbook()

    # Remove the default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    base_date = date(2025, 11, 1)

    # ------------------------------------------------------------------
    # Sheet 1: DISCOVERY — summary (not daily)
    # ------------------------------------------------------------------
    ws_discovery = wb.create_sheet("DISCOVERY")
    ws_discovery.append(["Overall Performance", "11/1/2025 - 11/30/2025"])
    ws_discovery.append(["Impressions", 15800])
    ws_discovery.append(["Members reached", 9200])

    # ------------------------------------------------------------------
    # Sheet 2: ENGAGEMENT — daily totals
    # ------------------------------------------------------------------
    ws_engagement = wb.create_sheet("ENGAGEMENT")
    ws_engagement.append(["Date", "Impressions", "Engagements"])
    for i in range(30):
        d = base_date + timedelta(days=i)
        impressions = 200 + (i % 7) * 50
        engagements = int(impressions * 0.03)
        ws_engagement.append([d.strftime("%m/%d/%Y"), impressions, engagements])

    # ------------------------------------------------------------------
    # Sheet 3: TOP POSTS — two side-by-side tables
    # ------------------------------------------------------------------
    ws_top = wb.create_sheet("TOP POSTS")
    ws_top.append(["Maximum of 50 posts available to include in this list"])
    ws_top.append([None])  # empty row 2

    # Header row (row 3): A-C for engagements, E-G for impressions
    ws_top.cell(row=3, column=1, value="Post URL")
    ws_top.cell(row=3, column=2, value="Post publish date")
    ws_top.cell(row=3, column=3, value="Engagements")
    ws_top.cell(row=3, column=5, value="Post URL")
    ws_top.cell(row=3, column=6, value="Post publish date")
    ws_top.cell(row=3, column=7, value="Impressions")

    # Data rows
    posts_data = [
        ("7400000000000000001", base_date, 180, 3200),
        ("7400000000000000002", base_date + timedelta(days=7), 155, 2800),
        ("7400000000000000003", base_date + timedelta(days=14), 210, 4500),
        ("7400000000000000004", base_date + timedelta(days=21), 280, 5100),
        ("7400000000000000005", base_date + timedelta(days=28), 120, 2200),
    ]

    # Sort by engagements for left table, by impressions for right table
    by_engagements = sorted(posts_data, key=lambda x: x[2], reverse=True)
    by_impressions = sorted(posts_data, key=lambda x: x[3], reverse=True)

    for row_idx in range(len(posts_data)):
        row_num = row_idx + 4  # start at row 4

        # Left table (engagements)
        if row_idx < len(by_engagements):
            aid, pdate, eng, _ = by_engagements[row_idx]
            ws_top.cell(row=row_num, column=1, value=f"https://www.linkedin.com/feed/update/urn:li:activity:{aid}")
            ws_top.cell(row=row_num, column=2, value=pdate.strftime("%m/%d/%Y"))
            ws_top.cell(row=row_num, column=3, value=eng)

        # Right table (impressions)
        if row_idx < len(by_impressions):
            aid, pdate, _, imp = by_impressions[row_idx]
            ws_top.cell(row=row_num, column=5, value=f"https://www.linkedin.com/feed/update/urn:li:activity:{aid}")
            ws_top.cell(row=row_num, column=6, value=pdate.strftime("%m/%d/%Y"))
            ws_top.cell(row=row_num, column=7, value=imp)

    # ------------------------------------------------------------------
    # Sheet 4: FOLLOWERS — total at top, header at row 3
    # ------------------------------------------------------------------
    ws_followers = wb.create_sheet("FOLLOWERS")
    ws_followers.append(["Total followers on 11/30/2025:", 520])
    ws_followers.append([None])  # empty row 2
    ws_followers.append(["Date", "New followers"])
    followers_total = 450
    for i in range(30):
        d = base_date + timedelta(days=i)
        new = i % 5 + 1
        ws_followers.append([d.strftime("%m/%d/%Y"), new])

    # ------------------------------------------------------------------
    # Sheet 5: DEMOGRAPHICS — "Top Demographics" header
    # ------------------------------------------------------------------
    ws_demo = wb.create_sheet("DEMOGRAPHICS")
    ws_demo.append(["Top Demographics", "Value", "Percentage"])
    demo_rows = [
        ("Industries", "Information Technology", 0.325),
        ("Industries", "Financial Services", 0.182),
        ("Industries", "Cybersecurity", 0.140),
        ("Industries", "Healthcare", 0.085),
        ("Industries", "Other", 0.268),
        ("Job titles", "Director", 0.220),
        ("Job titles", "Manager", 0.185),
        ("Job titles", "Individual Contributor", 0.300),
        ("Job titles", "Executive", 0.120),
        ("Job titles", "Other", 0.175),
        ("Seniority", "Senior", 0.400),
        ("Seniority", "Director", 0.250),
        ("Seniority", "Entry", 0.150),
        ("Seniority", "VP", 0.120),
        ("Seniority", "C-Suite", 0.080),
        ("Locations", "United States", 0.550),
        ("Locations", "Canada", 0.120),
        ("Locations", "United Kingdom", 0.095),
        ("Locations", "India", 0.080),
        ("Locations", "Other", 0.155),
    ]
    for category, value, pct in demo_rows:
        ws_demo.append([category, value, pct])

    wb.save(output_path)


@pytest.fixture(scope="session")
def sample_xlsx_path() -> Path:
    """Return the path to the pre-built synthetic sample .xlsx fixture.

    The file is generated once per session and stored in tests/fixtures/.
    """
    FIXTURES_DIR.mkdir(parents=True, exist_ok=True)
    output = FIXTURES_DIR / "sample_export.xlsx"
    _build_sample_xlsx(output)
    return output


@pytest.fixture
def sample_xlsx_bytes(sample_xlsx_path) -> bytes:
    """Return the raw bytes of the sample .xlsx file."""
    return sample_xlsx_path.read_bytes()
