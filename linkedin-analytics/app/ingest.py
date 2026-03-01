"""LinkedIn analytics export ingestion pipeline.

Parses XLS/XLSX exports from LinkedIn's creator analytics page,
deduplicates records, and loads them into the SQLite database.

Verified format (from real LinkedIn Premium export, Feb 2026):
  Sheet "DISCOVERY"    -> Summary row: "Impressions" + value, "Members reached" + value
  Sheet "ENGAGEMENT"   -> Daily totals: Date | Impressions | Engagements
  Sheet "TOP POSTS"    -> Two side-by-side tables: Engagements (A-C) and Impressions (E-G)
                          Each has: Post URL | Post publish date | metric value
  Sheet "FOLLOWERS"    -> Header row at row 3: Date | New followers. Total at B1.
  Sheet "DEMOGRAPHICS" -> Structured: Category (col A) | Value (col B) | Percentage (col C)
"""

import hashlib
import logging
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
from sqlalchemy.orm import Session

from app.models import DailyMetric, DemographicSnapshot, FollowerSnapshot, Post, PostDemographic, Upload

logger = logging.getLogger(__name__)

# Maximum file size: 50 MB
MAX_FILE_SIZE_BYTES = 50 * 1024 * 1024

# Allowed file extensions
ALLOWED_EXTENSIONS = {".xlsx", ".xls", ".csv"}

# Sheet names expected in the LinkedIn export
SHEET_DISCOVERY = "DISCOVERY"
SHEET_ENGAGEMENT = "ENGAGEMENT"
SHEET_TOP_POSTS = "TOP POSTS"
SHEET_FOLLOWERS = "FOLLOWERS"
SHEET_DEMOGRAPHICS = "DEMOGRAPHICS"


@dataclass
class ParsedExport:
    """Structured result from parsing a LinkedIn analytics export file."""

    posts: list[dict[str, Any]] = field(default_factory=list)
    daily_metrics: list[dict[str, Any]] = field(default_factory=list)
    follower_snapshots: list[dict[str, Any]] = field(default_factory=list)
    demographic_snapshots: list[dict[str, Any]] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


@dataclass
class ImportStats:
    """Statistics from a completed import operation."""

    posts_upserted: int = 0
    daily_metrics_upserted: int = 0
    follower_snapshots_upserted: int = 0
    demographic_snapshots_upserted: int = 0
    warnings: list[str] = field(default_factory=list)

    @property
    def total_records(self) -> int:
        return (
            self.posts_upserted
            + self.daily_metrics_upserted
            + self.follower_snapshots_upserted
            + self.demographic_snapshots_upserted
        )


class IngestError(Exception):
    """Raised when ingestion cannot proceed."""


class DuplicateFileError(IngestError):
    """Raised when the exact same file has already been imported."""


def compute_file_hash(file_path: Path) -> str:
    """Compute SHA256 hash of a file for deduplication.

    Args:
        file_path: Path to the file.

    Returns:
        Hex-encoded SHA256 digest string.
    """
    sha256 = hashlib.sha256()
    with open(file_path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            sha256.update(chunk)
    return sha256.hexdigest()


def validate_upload(file_path: Path) -> None:
    """Validate an uploaded file before parsing.

    Args:
        file_path: Path to the uploaded file.

    Raises:
        IngestError: If the file fails validation.
    """
    if not file_path.exists():
        raise IngestError(f"File not found: {file_path}")

    if file_path.stat().st_size == 0:
        raise IngestError("Uploaded file is empty.")

    if file_path.stat().st_size > MAX_FILE_SIZE_BYTES:
        raise IngestError(
            f"File exceeds maximum size of {MAX_FILE_SIZE_BYTES // (1024 * 1024)} MB."
        )

    suffix = file_path.suffix.lower()
    if suffix not in ALLOWED_EXTENSIONS:
        raise IngestError(
            f"Unsupported file type '{suffix}'. Allowed: {', '.join(sorted(ALLOWED_EXTENSIONS))}"
        )


def _safe_int(value: Any, default: int = 0) -> int:
    """Convert a value to int, returning default on failure."""
    if value is None:
        return default
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return default


def _safe_float(value: Any, default: float = 0.0) -> float:
    """Convert a value to float, returning default on failure."""
    if value is None:
        return default
    try:
        return float(value)
    except (ValueError, TypeError):
        return default


def _parse_date(value: Any) -> date | None:
    """Parse a date value from various formats."""
    if value is None:
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    return None


def _extract_activity_id(url: str) -> str | None:
    """Extract the LinkedIn activity ID from a post URL.

    Example: https://www.linkedin.com/feed/update/urn:li:activity:7432392249344274432
    Returns: "7432392249344274432"
    """
    if not url:
        return None
    match = re.search(r"urn:li:activity:(\d+)", str(url))
    return match.group(1) if match else None


def _load_workbook(file_path: Path) -> openpyxl.Workbook:
    """Load an Excel workbook using openpyxl with data_only mode.

    Must use read_only=False because LinkedIn exports have unreliable
    dimension metadata (max_col=1 in read_only mode).
    """
    return openpyxl.load_workbook(file_path, read_only=False, data_only=True)


def _get_sheet(wb: openpyxl.Workbook, name: str) -> Any | None:
    """Get a worksheet by name (case-insensitive)."""
    for sheet_name in wb.sheetnames:
        if sheet_name.strip().upper() == name.upper():
            return wb[sheet_name]
    return None


def _parse_discovery_sheet(ws: Any, warnings: list[str]) -> dict[str, int]:
    """Parse DISCOVERY sheet into summary metrics.

    Real format (verified):
      A1: "Overall Performance"   B1: "2/22/2026 - 2/28/2026"
      A2: "Impressions"           B2: 1491
      A3: "Members reached"       B3: 875
    """
    summary: dict[str, int] = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        label = row[0].value
        value = row[1].value if len(row) > 1 else None
        if label and value is not None:
            key = str(label).strip().upper()
            if key == "IMPRESSIONS":
                summary["impressions"] = _safe_int(value)
            elif key == "MEMBERS REACHED":
                summary["members_reached"] = _safe_int(value)
    return summary


def _parse_engagement_sheet(ws: Any, warnings: list[str]) -> list[dict]:
    """Parse ENGAGEMENT sheet into daily account-level metrics.

    Real format (verified):
      Row 1: Date | Impressions | Engagements
      Row 2+: date values with daily totals
    """
    records: list[dict] = []
    header_row = None

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        cells = [c.value for c in row]
        if not cells or cells[0] is None:
            continue

        # Find the header row
        if str(cells[0]).strip().upper() == "DATE":
            header_row = [str(c).strip().upper() if c else "" for c in cells]
            continue

        if header_row is None:
            continue

        row_date = _parse_date(cells[0])
        if not row_date:
            continue

        impressions = _safe_int(cells[1]) if len(cells) > 1 else 0
        engagements = _safe_int(cells[2]) if len(cells) > 2 else 0

        records.append({
            "metric_date": row_date,
            "post_id": None,
            "impressions": impressions,
            "engagements": engagements,
        })

    return records


def _parse_top_posts_sheet(ws: Any, warnings: list[str]) -> list[dict]:
    """Parse TOP POSTS sheet into post records.

    Real format (verified): two side-by-side tables.
      Row 1: "Maximum of 50 posts available..."
      Row 2: (empty)
      Row 3: Post URL | Post publish date | Engagements | (gap) | Post URL | Post publish date | Impressions
      Row 4+: data rows

    Left table (cols A-C): top posts by engagements
    Right table (cols E-G): top posts by impressions
    We merge them by post URL to get both metrics per post.
    """
    records_by_url: dict[str, dict[str, Any]] = {}
    data_started = False

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        cells = [c.value for c in row]

        # Find the header row (contains "Post URL")
        if not data_started:
            if cells[0] and str(cells[0]).strip().upper() == "POST URL":
                data_started = True
            continue

        # Left table: engagement data (cols A-C)
        url_left = str(cells[0]).strip() if cells[0] else ""
        if url_left and url_left.startswith("http"):
            activity_id = _extract_activity_id(url_left)
            pub_date = _parse_date(cells[1]) if len(cells) > 1 else None
            engagements = _safe_int(cells[2]) if len(cells) > 2 else 0

            if activity_id and pub_date:
                if activity_id not in records_by_url:
                    records_by_url[activity_id] = {
                        "linkedin_post_id": activity_id,
                        "post_url": url_left,
                        "post_date": pub_date,
                        "impressions": 0,
                        "engagements": engagements,
                    }
                else:
                    records_by_url[activity_id]["engagements"] = engagements

        # Right table: impressions data (cols E-G, index 4-6)
        if len(cells) > 4:
            url_right = str(cells[4]).strip() if cells[4] else ""
            if url_right and url_right.startswith("http"):
                activity_id = _extract_activity_id(url_right)
                pub_date = _parse_date(cells[5]) if len(cells) > 5 else None
                impressions = _safe_int(cells[6]) if len(cells) > 6 else 0

                if activity_id and pub_date:
                    if activity_id not in records_by_url:
                        records_by_url[activity_id] = {
                            "linkedin_post_id": activity_id,
                            "post_url": url_right,
                            "post_date": pub_date,
                            "impressions": impressions,
                            "engagements": 0,
                        }
                    else:
                        records_by_url[activity_id]["impressions"] = impressions

    # Convert to post records
    posts: list[dict] = []
    for record in records_by_url.values():
        impressions = record.get("impressions", 0)
        engagements = record.get("engagements", 0)
        engagement_rate = engagements / impressions if impressions > 0 else 0.0

        posts.append({
            "post_date": record["post_date"],
            "title": None,
            "linkedin_post_id": record["linkedin_post_id"],
            "impressions": impressions,
            "reactions": engagements,  # LinkedIn lumps all engagements together
            "comments": 0,
            "shares": 0,
            "clicks": 0,
            "members_reached": 0,
            "engagement_rate": engagement_rate,
            "post_type": None,
        })

    return posts


def _parse_followers_sheet(ws: Any, warnings: list[str]) -> list[dict]:
    """Parse FOLLOWERS sheet into follower snapshots.

    Real format (verified):
      Row 1: "Total followers on 2/28/2026:" | 1506
      Row 2: (empty)
      Row 3: "Date" | "New followers"
      Row 4+: date | new_followers_count
    """
    records: list[dict] = []
    total_followers: int = 0
    header_found = False

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        cells = [c.value for c in row]
        if not cells or cells[0] is None:
            continue

        label = str(cells[0]).strip()

        # Extract total followers from row 1
        if label.upper().startswith("TOTAL FOLLOWERS"):
            total_followers = _safe_int(cells[1]) if len(cells) > 1 else 0
            continue

        # Find the header row
        if label.upper() == "DATE":
            header_found = True
            continue

        if not header_found:
            continue

        row_date = _parse_date(cells[0])
        if not row_date:
            continue

        new_followers = _safe_int(cells[1]) if len(cells) > 1 else 0

        records.append({
            "snapshot_date": row_date,
            "total_followers": total_followers,
            "new_followers": new_followers,
        })

    if not records and total_followers > 0:
        warnings.append("FOLLOWERS sheet: no daily data rows found.")

    return records


def _parse_demographics_sheet(ws: Any, warnings: list[str]) -> list[dict]:
    """Parse DEMOGRAPHICS sheet into demographic snapshots.

    Real format (verified):
      Row 1: "Top Demographics" | "Value" | "Percentage"
      Row 2+: category (col A) | value (col B) | percentage as float (col C)
      Categories repeat per row: "Job titles", "Locations", "Industries",
      "Seniority", "Company size", "Companies"
    """
    records: list[dict] = []
    snapshot_date = date.today()
    header_found = False

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        cells = [c.value for c in row]
        if not cells or cells[0] is None:
            continue

        label = str(cells[0]).strip()

        # Skip the header row
        if label.upper() in ("TOP DEMOGRAPHICS", "CATEGORY"):
            header_found = True
            continue

        if not header_found:
            continue

        category = label.lower()
        value = str(cells[1]).strip() if len(cells) > 1 and cells[1] else ""
        pct = _safe_float(cells[2]) if len(cells) > 2 else 0.0

        if not value:
            continue

        records.append({
            "snapshot_date": snapshot_date,
            "category": category,
            "value": value,
            "percentage": pct,
        })

    if not records:
        warnings.append("DEMOGRAPHICS sheet: could not parse any demographic records.")

    return records


def parse_linkedin_export(file_path: Path) -> ParsedExport:
    """Parse a LinkedIn analytics export file into structured data.

    Args:
        file_path: Path to the .xlsx, .xls, or .csv file.

    Returns:
        ParsedExport dataclass containing parsed records and any warnings.

    Raises:
        IngestError: If the file cannot be parsed.
    """
    validate_upload(file_path)

    result = ParsedExport()

    # CSV files are accepted but cannot be parsed as LinkedIn exports
    # (LinkedIn always exports .xlsx). Return empty result with a warning.
    if file_path.suffix.lower() == ".csv":
        result.warnings.append("CSV files are accepted but LinkedIn exports are .xlsx. No data parsed.")
        return result

    try:
        wb = _load_workbook(file_path)
    except Exception as exc:
        raise IngestError(f"Failed to read file '{file_path.name}': {exc}") from exc

    sheet_names = [s.strip().upper() for s in wb.sheetnames]
    logger.info("Loaded sheets: %s", sorted(sheet_names))

    if not sheet_names:
        wb.close()
        raise IngestError("No sheets found in file.")

    try:
        # DISCOVERY sheet -> summary metrics (stored as info, not daily rows)
        ws = _get_sheet(wb, SHEET_DISCOVERY)
        if ws:
            summary = _parse_discovery_sheet(ws, result.warnings)
            if summary:
                logger.info("Discovery summary: %s", summary)
        else:
            result.warnings.append(f"Sheet '{SHEET_DISCOVERY}' not found.")

        # ENGAGEMENT sheet -> daily account-level metrics
        ws = _get_sheet(wb, SHEET_ENGAGEMENT)
        if ws:
            result.daily_metrics.extend(
                _parse_engagement_sheet(ws, result.warnings)
            )
        else:
            result.warnings.append(f"Sheet '{SHEET_ENGAGEMENT}' not found.")

        # TOP POSTS sheet -> individual post records with URLs
        ws = _get_sheet(wb, SHEET_TOP_POSTS)
        if ws:
            result.posts.extend(
                _parse_top_posts_sheet(ws, result.warnings)
            )
        else:
            result.warnings.append(f"Sheet '{SHEET_TOP_POSTS}' not found.")

        # FOLLOWERS sheet -> follower snapshots
        ws = _get_sheet(wb, SHEET_FOLLOWERS)
        if ws:
            result.follower_snapshots.extend(
                _parse_followers_sheet(ws, result.warnings)
            )
        else:
            result.warnings.append(f"Sheet '{SHEET_FOLLOWERS}' not found.")

        # DEMOGRAPHICS sheet -> demographic snapshots
        ws = _get_sheet(wb, SHEET_DEMOGRAPHICS)
        if ws:
            result.demographic_snapshots.extend(
                _parse_demographics_sheet(ws, result.warnings)
            )
        else:
            result.warnings.append(f"Sheet '{SHEET_DEMOGRAPHICS}' not found.")
    finally:
        wb.close()

    logger.info(
        "Parsed: %d posts, %d daily metrics, %d follower snapshots, %d demographic records",
        len(result.posts),
        len(result.daily_metrics),
        len(result.follower_snapshots),
        len(result.demographic_snapshots),
    )

    return result


def _upsert_post(session: Session, record: dict[str, Any]) -> Post:
    """Insert or update a Post record using UPSERT semantics.

    The higher value wins for cumulative metrics when updating an existing post.
    """
    existing: Post | None = None

    # Try LinkedIn post ID first (most reliable)
    if record.get("linkedin_post_id"):
        existing = session.query(Post).filter_by(
            linkedin_post_id=record["linkedin_post_id"]
        ).first()

    # Fall back to composite key: post_date + title[:100]
    if not existing and record.get("post_date") and record.get("title"):
        existing = (
            session.query(Post)
            .filter_by(post_date=record["post_date"], title=record["title"])
            .first()
        )

    # Fall back to date-only match when title is None
    if not existing and record.get("post_date") and not record.get("title"):
        existing = (
            session.query(Post)
            .filter_by(post_date=record["post_date"], title=None)
            .first()
        )

    if existing:
        # Higher value wins for cumulative metrics
        existing.impressions = max(existing.impressions or 0, record.get("impressions", 0))
        existing.members_reached = max(existing.members_reached or 0, record.get("members_reached", 0))
        existing.reactions = max(existing.reactions or 0, record.get("reactions", 0))
        existing.comments = max(existing.comments or 0, record.get("comments", 0))
        existing.shares = max(existing.shares or 0, record.get("shares", 0))
        existing.clicks = max(existing.clicks or 0, record.get("clicks", 0))
        if record.get("linkedin_post_id") and not existing.linkedin_post_id:
            existing.linkedin_post_id = record["linkedin_post_id"]
        if record.get("post_url") and not existing.post_url:
            existing.post_url = record["post_url"]
        if record.get("post_type") and not existing.post_type:
            existing.post_type = record["post_type"]
        existing.recalculate_engagement_rate()
        # Transition status: if the post was published via the API and now has analytics, mark as linked
        if existing.status == "published" and existing.content:
            existing.status = "analytics_linked"
        return existing

    post = Post(
        linkedin_post_id=record.get("linkedin_post_id"),
        post_url=record.get("post_url"),
        title=record.get("title"),
        post_date=record["post_date"],
        post_type=record.get("post_type"),
        impressions=record.get("impressions", 0),
        members_reached=record.get("members_reached", 0),
        reactions=record.get("reactions", 0),
        comments=record.get("comments", 0),
        shares=record.get("shares", 0),
        clicks=record.get("clicks", 0),
        engagement_rate=record.get("engagement_rate", 0.0),
    )
    session.add(post)
    return post


def _upsert_daily_metric(session: Session, record: dict[str, Any]) -> None:
    """Insert or update a DailyMetric record."""
    post_id = record.get("post_id")
    if post_id is None:
        # SQLite treats NULL != NULL, so filter_by(post_id=None) won't match.
        # Use explicit IS NULL filter for account-level metrics.
        from sqlalchemy import null
        existing = (
            session.query(DailyMetric)
            .filter(
                DailyMetric.post_id.is_(None),
                DailyMetric.metric_date == record["metric_date"],
            )
            .first()
        )
    else:
        existing = (
            session.query(DailyMetric)
            .filter_by(post_id=post_id, metric_date=record["metric_date"])
            .first()
        )
    if existing:
        existing.impressions = max(existing.impressions or 0, record.get("impressions", 0))
        existing.members_reached = max(existing.members_reached or 0, record.get("members_reached", 0))
        existing.reactions = max(existing.reactions or 0, record.get("reactions", 0))
        existing.comments = max(existing.comments or 0, record.get("comments", 0))
        existing.shares = max(existing.shares or 0, record.get("shares", 0))
        existing.clicks = max(existing.clicks or 0, record.get("clicks", 0))
        return

    metric = DailyMetric(
        post_id=record.get("post_id"),
        metric_date=record["metric_date"],
        impressions=record.get("impressions", 0),
        members_reached=record.get("members_reached", 0),
        reactions=record.get("reactions", 0),
        comments=record.get("comments", 0),
        shares=record.get("shares", 0),
        clicks=record.get("clicks", 0),
    )
    session.add(metric)


def _upsert_follower_snapshot(session: Session, record: dict[str, Any]) -> None:
    """Insert or update a FollowerSnapshot record."""
    existing = (
        session.query(FollowerSnapshot)
        .filter_by(snapshot_date=record["snapshot_date"])
        .first()
    )
    if existing:
        existing.total_followers = record["total_followers"]
        existing.new_followers = record.get("new_followers", 0)
        return

    snapshot = FollowerSnapshot(
        snapshot_date=record["snapshot_date"],
        total_followers=record["total_followers"],
        new_followers=record.get("new_followers", 0),
    )
    session.add(snapshot)


def _upsert_demographic_snapshot(session: Session, record: dict[str, Any]) -> None:
    """Insert or update a DemographicSnapshot record."""
    existing = (
        session.query(DemographicSnapshot)
        .filter_by(
            snapshot_date=record["snapshot_date"],
            category=record["category"],
            value=record["value"],
        )
        .first()
    )
    if existing:
        existing.percentage = record["percentage"]
        return

    snapshot = DemographicSnapshot(
        snapshot_date=record["snapshot_date"],
        category=record["category"],
        value=record["value"],
        percentage=record["percentage"],
    )
    session.add(snapshot)


def load_to_db(session: Session, parsed: ParsedExport) -> ImportStats:
    """Load parsed export data into the database.

    Uses UPSERT semantics: higher value wins for cumulative metrics.

    Args:
        session: SQLAlchemy session.
        parsed: ParsedExport from parse_linkedin_export().

    Returns:
        ImportStats with counts of upserted records.
    """
    stats = ImportStats(warnings=list(parsed.warnings))

    # Upsert posts
    for record in parsed.posts:
        try:
            _upsert_post(session, record)
            stats.posts_upserted += 1
        except Exception as exc:
            msg = f"Failed to upsert post (date={record.get('post_date')}): {exc}"
            logger.warning(msg)
            stats.warnings.append(msg)

    session.flush()  # Ensure post IDs are available for daily metrics

    # Upsert daily account-level metrics (post_id=None)
    for record in parsed.daily_metrics:
        try:
            _upsert_daily_metric(session, record)
            stats.daily_metrics_upserted += 1
        except Exception as exc:
            msg = f"Failed to upsert daily metric (date={record.get('metric_date')}): {exc}"
            logger.warning(msg)
            stats.warnings.append(msg)

    # Upsert follower snapshots
    for record in parsed.follower_snapshots:
        try:
            _upsert_follower_snapshot(session, record)
            stats.follower_snapshots_upserted += 1
        except Exception as exc:
            msg = f"Failed to upsert follower snapshot (date={record.get('snapshot_date')}): {exc}"
            logger.warning(msg)
            stats.warnings.append(msg)

    # Upsert demographic snapshots
    for record in parsed.demographic_snapshots:
        try:
            _upsert_demographic_snapshot(session, record)
            stats.demographic_snapshots_upserted += 1
        except Exception as exc:
            msg = f"Failed to upsert demographic (category={record.get('category')}): {exc}"
            logger.warning(msg)
            stats.warnings.append(msg)

    session.commit()

    logger.info(
        "Import complete: %d posts, %d daily metrics, %d follower snapshots, %d demographics",
        stats.posts_upserted,
        stats.daily_metrics_upserted,
        stats.follower_snapshots_upserted,
        stats.demographic_snapshots_upserted,
    )

    return stats


# ---------------------------------------------------------------------------
# Per-post XLSX export detection and parsing
# ---------------------------------------------------------------------------


def _detect_xlsx_format(wb: openpyxl.Workbook) -> str:
    """Detect XLSX export format by sheet names.

    Returns:
        "per_post" if PERFORMANCE and TOP DEMOGRAPHICS sheets are present.
        "aggregate" if DISCOVERY and ENGAGEMENT sheets are present.
        "unknown" otherwise.
    """
    sheet_names = {s.strip().lower() for s in wb.sheetnames}
    if "performance" in sheet_names and "top demographics" in sheet_names:
        return "per_post"
    if "discovery" in sheet_names and "engagement" in sheet_names:
        return "aggregate"
    return "unknown"


def _parse_per_post_performance(ws: Any) -> dict[str, str]:
    """Parse key-value pairs from the PERFORMANCE sheet.

    The PERFORMANCE sheet uses a key-value layout (not tabular):
    - Row 1: Post URL = https://...urn:li:share:1234
    - Row 2: Post Date = Feb 25, 2026
    - Row 3: Post Publish Time = 11:53 AM
    - Row 5+: Metric = value pairs (Impressions, Reactions, etc.)

    Values in column B are strings (may include commas in numbers like "1,316").
    """
    data: dict[str, str] = {}
    for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
        if row[0] and row[1] is not None:
            key = str(row[0]).strip()
            val = str(row[1]).strip()
            data[key] = val
    return data


def _parse_int_with_commas(s: str) -> int:
    """Parse integer string that may contain commas. Returns 0 on failure."""
    try:
        return int(str(s).replace(",", ""))
    except (ValueError, AttributeError):
        return 0


def _parse_post_hour(time_str: str) -> int | None:
    """Parse post hour from LinkedIn time format (e.g., '11:53 AM').

    Returns hour in 24-hour format (0-23), or None on failure.
    """
    try:
        from datetime import datetime as _dt
        t = _dt.strptime(time_str.strip(), "%I:%M %p")
        return t.hour
    except (ValueError, AttributeError):
        return None


def _extract_urn_from_url(url: str) -> str | None:
    """Extract the numeric share ID from a LinkedIn post URL.

    Handles URLs containing urn:li:share:{id} or urn:li:activity:{id}.
    """
    match = re.search(r"urn:li:(?:share|activity):(\d+)", url)
    return match.group(1) if match else None


def _parse_per_post_demographics(ws: Any) -> list[dict[str, Any]]:
    """Parse the TOP DEMOGRAPHICS sheet.

    Tabular format: Category | Value | Percentage
    Categories: "Company size", "Job title", "Location", "Company"
    Percentages are floats (0.31) or strings ("< 1%").
    """
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
        if not row[0] or row[1] is None:
            continue
        category = str(row[0]).strip().lower().replace(" ", "_")
        value = str(row[1]).strip()
        pct_raw = row[2]
        if isinstance(pct_raw, (int, float)):
            percentage = float(pct_raw)
        elif isinstance(pct_raw, str) and "<" in pct_raw:
            percentage = 0.005  # "< 1%" stored as 0.5%
        else:
            try:
                percentage = float(str(pct_raw).strip().rstrip("%")) / 100
            except (ValueError, AttributeError):
                percentage = 0.0
        rows.append({
            "category": category,
            "value": value,
            "percentage": percentage,
        })
    return rows


def ingest_per_post_xlsx(session: Session, wb: openpyxl.Workbook) -> dict[str, Any]:
    """Ingest a per-post XLSX export.

    Extracts metrics, post_hour, linkedin_post_id, and demographics
    from the per-post export format.

    Args:
        session: SQLAlchemy session.
        wb: Open openpyxl workbook (must contain PERFORMANCE and TOP DEMOGRAPHICS sheets).

    Returns:
        Dict with import results: post_id, linkedin_post_id, metrics_updated, demographics_imported.
    """
    # Find sheets case-insensitively
    perf_ws = None
    demo_ws = None
    for sheet_name in wb.sheetnames:
        upper = sheet_name.strip().upper()
        if upper == "PERFORMANCE":
            perf_ws = wb[sheet_name]
        elif upper == "TOP DEMOGRAPHICS":
            demo_ws = wb[sheet_name]

    if perf_ws is None:
        raise IngestError("Per-post XLSX is missing the PERFORMANCE sheet.")
    if demo_ws is None:
        raise IngestError("Per-post XLSX is missing the TOP DEMOGRAPHICS sheet.")

    # Parse performance data
    perf = _parse_per_post_performance(perf_ws)

    post_url = perf.get("Post URL", "")
    linkedin_post_id = _extract_urn_from_url(post_url)
    post_hour = _parse_post_hour(perf.get("Post Publish Time", ""))

    metrics = {
        "impressions": _parse_int_with_commas(perf.get("Impressions", "0")),
        "members_reached": _parse_int_with_commas(perf.get("Members reached", "0")),
        "reactions": _parse_int_with_commas(perf.get("Reactions", "0")),
        "comments": _parse_int_with_commas(perf.get("Comments", "0")),
        "reposts": _parse_int_with_commas(perf.get("Reposts", "0")),
        "saves": _parse_int_with_commas(perf.get("Saves", "0")),
        "sends": _parse_int_with_commas(perf.get("Sends on LinkedIn", "0")),
        "profile_views": _parse_int_with_commas(
            perf.get("Profile viewers from this post", "0")
        ),
        "followers_gained": _parse_int_with_commas(
            perf.get("Followers gained from this post", "0")
        ),
    }

    # Parse post date early so it's available for fallback matching.
    post_date_str = perf.get("Post Date", "")
    try:
        from datetime import datetime as _dt
        post_date = _dt.strptime(post_date_str, "%b %d, %Y").date()
    except (ValueError, AttributeError):
        from datetime import date as _d
        post_date = _d.today()

    # Find or create the post.
    # Primary match: by share/activity URN extracted from per-post XLSX.
    # Fallback: match by post_date, since aggregate and per-post XLSX exports
    # use different URN types (urn:li:activity vs urn:li:share) with different
    # numeric IDs for the same post.
    existing_post: Post | None = None
    if linkedin_post_id:
        existing_post = (
            session.query(Post)
            .filter(Post.linkedin_post_id == linkedin_post_id)
            .first()
        )
    if not existing_post:
        existing_post = (
            session.query(Post)
            .filter(Post.post_date == post_date)
            .first()
        )

    if existing_post:
        post = existing_post
        # Per-post export has more granular data; overwrite with exact values
        for key, val in metrics.items():
            if hasattr(post, key):
                setattr(post, key, val)
        if post_hour is not None:
            post.post_hour = post_hour
        # Store the share URL if we don't have one yet.
        if post_url and not post.post_url:
            post.post_url = post_url
        post.recalculate_engagement_rate()
        # Transition status if applicable
        if post.status == "published" and post.content:
            post.status = "analytics_linked"
    else:

        post = Post(
            linkedin_post_id=linkedin_post_id,
            post_url=post_url if post_url else None,
            post_date=post_date,
            post_hour=post_hour,
            impressions=metrics.get("impressions", 0),
            members_reached=metrics.get("members_reached", 0),
            reactions=metrics.get("reactions", 0),
            comments=metrics.get("comments", 0),
            reposts=metrics.get("reposts", 0),
            saves=metrics.get("saves", 0),
            sends=metrics.get("sends", 0),
            profile_views=metrics.get("profile_views", 0),
            followers_gained=metrics.get("followers_gained", 0),
        )
        post.recalculate_engagement_rate()
        session.add(post)
        session.flush()  # Get post.id for demographics FK

    # Parse and store demographics
    demo_rows = _parse_per_post_demographics(demo_ws)
    demo_count = 0
    for row in demo_rows:
        existing_demo = (
            session.query(PostDemographic)
            .filter(
                PostDemographic.post_id == post.id,
                PostDemographic.category == row["category"],
                PostDemographic.value == row["value"],
            )
            .first()
        )
        if existing_demo:
            existing_demo.percentage = row["percentage"]
        else:
            session.add(
                PostDemographic(
                    post_id=post.id,
                    category=row["category"],
                    value=row["value"],
                    percentage=row["percentage"],
                )
            )
            demo_count += 1

    session.commit()

    return {
        "post_id": post.id,
        "linkedin_post_id": linkedin_post_id,
        "metrics_updated": True,
        "demographics_imported": demo_count,
    }


def ingest_file(
    session: Session,
    file_path: Path,
    original_filename: str,
) -> tuple[Upload, ImportStats]:
    """Full ingestion pipeline: validate, deduplicate, parse, and load.

    Args:
        session: SQLAlchemy session.
        file_path: Path to the saved upload file.
        original_filename: Original filename from the HTTP upload.

    Returns:
        Tuple of (Upload record, ImportStats).

    Raises:
        DuplicateFileError: If the file has already been imported.
        IngestError: If ingestion fails for any other reason.
    """
    # File-level deduplication via SHA256
    file_hash = compute_file_hash(file_path)
    existing_upload = session.query(Upload).filter_by(file_hash=file_hash).first()
    if existing_upload:
        raise DuplicateFileError(
            f"File '{original_filename}' has already been imported "
            f"(uploaded as '{existing_upload.filename}' on {existing_upload.upload_date})."
        )

    # Auto-detect format for XLSX files
    if file_path.suffix.lower() in (".xlsx", ".xls"):
        try:
            wb = _load_workbook(file_path)
            fmt = _detect_xlsx_format(wb)
            wb.close()
        except Exception:
            fmt = "aggregate"  # Fall through to standard parser on error

        if fmt == "per_post":
            try:
                wb = _load_workbook(file_path)
                per_post_result = ingest_per_post_xlsx(session, wb)
                wb.close()
            except Exception as exc:
                raise IngestError(f"Per-post XLSX ingest failed: {exc}") from exc

            upload = Upload(
                filename=original_filename,
                file_hash=file_hash,
                records_imported=1,
                status="completed",
            )
            session.add(upload)
            session.commit()

            # Return a minimal ImportStats compatible result
            stats = ImportStats(posts_upserted=1)
            return upload, stats

    parsed = parse_linkedin_export(file_path)
    stats = load_to_db(session, parsed)

    upload = Upload(
        filename=original_filename,
        file_hash=file_hash,
        records_imported=stats.total_records,
        status="completed",
    )
    session.add(upload)
    session.commit()

    return upload, stats
